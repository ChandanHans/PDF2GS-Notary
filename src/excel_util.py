import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_dropdown(df, file_path):
    color_options = ['à envoyer', 'draft', 'envoyé', 'pas trouvé']
    color_codes = ['#ff8e8e',"#ffeeb0", '#b2ffaf', '#daeef3']
    
    workbook = xlsxwriter.Workbook(file_path, {'nan_inf_to_errors': True})
    worksheet = workbook.add_worksheet()
    
    # Write column headers
    worksheet.write_row('A1', df.columns)
    
    # Write data rows
    for idx, row in df.iterrows():
        worksheet.write_row(idx + 1, 0, row)
    
    col = 10
    for row in range(1, len(df) + 1):
        validation = {
            'validate': 'list',
            'source': color_options,
        }
        
        worksheet.data_validation(row, col, row, col, validation)
        worksheet.write(row, col, color_options[0])
        
        for i, option in enumerate(color_options):
            color_range = f'K{row + 1}'
            
            worksheet.conditional_format(
                color_range,
                {
                    'type': 'formula',
                    'criteria': f'=$K${row + 1}="{option}"',
                    'format': workbook.add_format({'bg_color': color_codes[i]}),
                }
            )
    
    workbook.close()

def hide_columns(worksheet, column_list):
    for col_letter in column_list:
        column = worksheet.column_dimensions[col_letter]
        column.hidden = True

def save_table(df, file_path):
    create_dropdown(df, file_path)
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    columns_to_hide = ['B', 'C', 'D', 'H', 'G', 'J']
    hide_columns(worksheet, columns_to_hide)
    
    table_style = TableStyleInfo(
        name='TableStyleMedium6',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table = Table(displayName='MyTable', ref=worksheet.dimensions, tableStyleInfo=table_style)
    worksheet.add_table(table)
    for col_num, _ in enumerate(df.columns, start = 0):
        worksheet.column_dimensions[xlsxwriter.utility.xl_col_to_name(col_num)].width = 30
    
    workbook.save(file_path)